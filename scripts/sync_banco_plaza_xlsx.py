#!/usr/bin/env python3
"""
Banco Plaza XLSX Sync
======================
Reads submitted responses from prod Odoo (ir.config_parameter 'banco_plaza.submissions')
and merges them into the local XLSX file.

Run on demand after employees have submitted the form.

Usage:
    python3 scripts/sync_banco_plaza_xlsx.py          # preview — print changes, no write
    python3 scripts/sync_banco_plaza_xlsx.py --apply  # write updated XLSX
"""

import argparse
import json
import logging
import sys
import xmlrpc.client

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

XLSX_PATH         = '/home/ftpuser/odoo-dev/Plantilla-Empleados-UEIPAB-FILLED(1).xlsx'
PROD_CFG          = '/opt/odoo-dev/config/production.json'
PARAM_SUBMISSIONS = 'banco_plaza.submissions'
DATA_ROW_START    = 8


def _connect():
    cfg = json.load(open(PROD_CFG))['production']['xmlrpc']
    url, db, user, key = cfg['url'], cfg['db'], cfg['user'], cfg['api_key']
    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
    uid = common.authenticate(db, user, key, {})
    if not uid:
        raise RuntimeError('XML-RPC auth failed')
    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
    return models, db, uid, key


def _get_submissions(models, db, uid, key) -> dict:
    result = models.execute_kw(db, uid, key,
                               'ir.config_parameter', 'search_read',
                               [[['key', '=', PARAM_SUBMISSIONS]]],
                               {'fields': ['value'], 'limit': 1})
    if not result:
        return {}
    return json.loads(result[0]['value'])


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--apply', action='store_true',
                    help='Write changes to XLSX (default: preview only)')
    args = ap.parse_args()

    log.info("Connecting to prod Odoo...")
    models, db, uid, key = _connect()

    submissions = _get_submissions(models, db, uid, key)
    log.info("Found %d submission(s) in prod Odoo.", len(submissions))

    if not submissions:
        log.info("Nothing to sync.")
        return

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    updates = 0
    for row in range(DATA_ROW_START, ws.max_row + 1):
        email_cell = ws.cell(row, 13).value   # column M
        if not email_cell:
            continue
        email = email_cell.strip().lower()
        sub = submissions.get(email)
        if not sub:
            continue

        nombre = (ws.cell(row, 3).value or '') + ' ' + (ws.cell(row, 5).value or '')
        changes = []

        # D — Segundo Nombre
        if sub.get('segundo_nombre') is not None:
            old = ws.cell(row, 4).value
            new = sub['segundo_nombre'] or None
            if old != new:
                changes.append(f'D: {old!r} → {new!r}')
                if args.apply:
                    ws.cell(row, 4).value = new

        # F — Segundo Apellido
        if sub.get('segundo_apellido') is not None:
            old = ws.cell(row, 6).value
            new = sub['segundo_apellido'] or None
            if old != new:
                changes.append(f'F: {old!r} → {new!r}')
                if args.apply:
                    ws.cell(row, 6).value = new

        # O — Operadora
        if sub.get('operadora'):
            old = ws.cell(row, 15).value
            new = int(sub['operadora'])
            if old != new:
                changes.append(f'O: {old!r} → {new!r}')
                if args.apply:
                    ws.cell(row, 15).value = new

        # P — Número
        if sub.get('numero'):
            old = ws.cell(row, 16).value
            new = int(sub['numero'])
            if old != new:
                changes.append(f'P: {old!r} → {new!r}')
                if args.apply:
                    ws.cell(row, 16).value = new

        if changes:
            count = sub.get('submission_count', 1)
            ts    = sub.get('submitted_at', '')[:19]
            log.info("  Row %-3d %-25s  [submit #%d @ %s]  %s",
                     row, nombre.strip(), count, ts, ' | '.join(changes))
            updates += 1
        else:
            log.info("  Row %-3d %-25s  no changes", row, nombre.strip())

    if updates == 0:
        log.info("All rows already up to date.")
        return

    if args.apply:
        wb.save(XLSX_PATH)
        log.info("✅ XLSX updated: %d row(s) modified → %s", updates, XLSX_PATH)
    else:
        log.info("Preview: %d row(s) would be updated. Run with --apply to write.", updates)


if __name__ == '__main__':
    main()
