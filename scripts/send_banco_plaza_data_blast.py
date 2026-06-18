#!/usr/bin/env python3
"""
Banco Plaza Employee Data Blast
================================
Sends a personalised email to each employee in the Plantilla-Empleados XLSX
asking them to confirm/provide:
  D  – Segundo Nombre
  F  – Segundo Apellido
  O  – Operadora  (3 digits: 412 / 414 / 416 / 424 / 426)
  P  – Número Telefónico (7 digits)

FROM     : UEIPAB Recursos Humanos <recursoshumanos@ueipab.edu.ve>
REPLY-TO : recursoshumanos@ueipab.edu.ve
CC       : recursoshumanos@ueipab.edu.ve

Usage:
    python3 scripts/send_banco_plaza_data_blast.py            # dry-run (CEO only)
    python3 scripts/send_banco_plaza_data_blast.py --preview  # CEO only, real HTML
    python3 scripts/send_banco_plaza_data_blast.py --live     # full blast
"""

import argparse
import hashlib
import hmac as _hmac
import json
import logging
import re
import secrets
import sys
import xmlrpc.client
from datetime import datetime

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Config ─────────────────────────────────────────────────────────────────────

XLSX_PATH    = '/home/ftpuser/odoo-dev/Plantilla-Empleados-UEIPAB-FILLED(1).xlsx'
PROD_CFG     = '/opt/odoo-dev/config/production.json'
ODOO_URL     = 'https://odoo.ueipab.edu.ve'
LOGO_URL     = f'{ODOO_URL}/web/image/res.company/1/logo'
FORM_BASE    = f'{ODOO_URL}/banco-plaza-form'

EMAIL_FROM   = 'UEIPAB Recursos Humanos <recursoshumanos@ueipab.edu.ve>'
CC_EMAIL     = 'recursoshumanos@ueipab.edu.ve'
REPLY_TO     = 'recursoshumanos@ueipab.edu.ve'
SUBJECT      = 'Banco Plaza — Verificación de datos de nómina | UEIPAB'
CEO_EMAIL    = 'gustavo.perdomo@ueipab.edu.ve'

PARAM_SECRET    = 'banco_plaza.form_secret'
PARAM_EMPLOYEES = 'banco_plaza.employees'
PARAM_OPEN      = 'banco_plaza.campaign_open'

MAIL_QUEUE_CRON_ID = 3
DATA_ROW_START     = 8   # first employee row in XLSX
VALID_OPS          = {412, 414, 416, 421, 424, 426}
BACKUP_NUM         = '4148321963'  # backup WA — never a personal phone

# Values we can infer with confidence from full Odoo employee names (4-part)
ODOO_NAME_OVERRIDE = {
    'gustavo.perdomo@ueipab.edu.ve':  ('JOSE',     'MATA'),
    'alejandra.lopez@ueipab.edu.ve':  ('CRISTINA', 'SAYAGO'),
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def _load_prod_cfg():
    cfg = json.load(open(PROD_CFG))['production']['xmlrpc']
    return cfg['url'], cfg['db'], cfg['user'], cfg['api_key']


def _connect():
    url, db, user, key = _load_prod_cfg()
    common = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/common')
    uid = common.authenticate(db, user, key, {})
    if not uid:
        raise RuntimeError('XML-RPC auth failed')
    models = xmlrpc.client.ServerProxy(f'{url}/xmlrpc/2/object')
    return models, db, uid, key


def call(models, db, uid, key, model, method, args=None, kw=None):
    return models.execute_kw(db, uid, key, model, method, args or [[]], kw or {})


# ── Token helpers ──────────────────────────────────────────────────────────────

def _compute_token(secret: str, email: str) -> str:
    return _hmac.new(
        secret.encode(), email.lower().encode(), hashlib.sha256
    ).hexdigest()[:24]


def _init_odoo_params(employees: list, models, db, uid, key):
    """Ensure form_secret exists in prod Odoo and store enriched employee list."""
    # Get or create secret
    result = call(models, db, uid, key, 'ir.config_parameter', 'search_read',
                  [[['key', '=', PARAM_SECRET]]], {'fields': ['value'], 'limit': 1})
    if result:
        secret = result[0]['value']
        log.info("Token secret: existing (first 8: %s...)", secret[:8])
    else:
        secret = secrets.token_hex(32)
        models.execute_kw(db, uid, key, 'ir.config_parameter', 'set_param',
                          [PARAM_SECRET, secret])
        log.info("Token secret: created (first 8: %s...)", secret[:8])

    # Store employee list (serializable subset of each emp dict)
    emp_list = [{k: v for k, v in e.items()
                 if k in ('email', 'primer_nombre', 'segundo_nombre',
                          'primer_apellido', 'segundo_apellido', 'cedula',
                          'dob', 'estado_civil', 'sexo', 'operadora', 'numero')}
                for e in employees]
    models.execute_kw(db, uid, key, 'ir.config_parameter', 'set_param',
                      [PARAM_EMPLOYEES, json.dumps(emp_list, ensure_ascii=False)])

    # Mark campaign open
    models.execute_kw(db, uid, key, 'ir.config_parameter', 'set_param',
                      [PARAM_OPEN, 'True'])
    log.info("Stored %d employees in prod Odoo params. Campaign open=True.", len(emp_list))

    return secret


def _parse_odoo_phone(mobile: str):
    """Return (operadora_int, numero_7_str) or (None, None)."""
    if not mobile:
        return None, None
    digits = re.sub(r'[^\d]', '', mobile)
    if digits.startswith('58') and len(digits) >= 12:
        digits = digits[2:]
    if len(digits) == 10:
        op = int(digits[:3])
        num = digits[3:]          # 7 digits, preserves leading zero
        if op in VALID_OPS:
            return op, num
    return None, None


def _is_valid_xlsx_phone(op_raw, num_raw) -> tuple:
    """Validate raw XLSX O/P values. Returns (op_int, num_str) or (None, None)."""
    if not op_raw or not num_raw:
        return None, None
    try:
        op_i  = int(op_raw)
        num_s = str(int(num_raw)).zfill(7)   # ensure 7 digits
        combined = f"{op_i}{num_s}"
        if op_i in VALID_OPS and len(num_s) == 7 and combined != BACKUP_NUM:
            return op_i, num_s
    except (ValueError, TypeError):
        pass
    return None, None


# ── XLSX reader ────────────────────────────────────────────────────────────────

def _load_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    employees = []
    for row in range(DATA_ROW_START, ws.max_row + 1):
        v = [ws.cell(row, c).value for c in range(1, 17)]
        # Skip footer/empty rows
        if not v[0] or v[0] in ('FOR-NE-004-167 (01-2025) V0',):
            continue
        email = (v[12] or '').strip().lower()
        if not email or '@' not in email:
            continue

        dob = v[9]
        if isinstance(dob, datetime):
            dob_str = dob.strftime('%d/%m/%Y')
        else:
            dob_str = str(dob) if dob else ''

        op, num = _is_valid_xlsx_phone(v[14], v[15])

        employees.append({
            'row':            row,
            'cedula':         v[1],
            'primer_nombre':  (v[2] or '').strip(),
            'segundo_nombre': (v[3] or '').strip() or None,   # D
            'primer_apellido':(v[4] or '').strip(),
            'segundo_apellido':(v[5] or '').strip() or None,  # F
            'dob':            dob_str,
            'estado_civil':   v[10] or '',
            'sexo':           v[11] or '',
            'email':          email,
            'operadora':      op,     # O
            'numero':         num,    # P
            'phone_source':   'xlsx' if op else None,
            'seg_nom_source': 'xlsx' if v[3] else None,
            'seg_ap_source':  'xlsx' if v[5] else None,
        })
    log.info("XLSX: %d employees loaded", len(employees))
    return employees


# ── Odoo enrichment ────────────────────────────────────────────────────────────

def _enrich_from_odoo(employees, models, db, uid, key):
    odoo_emps = call(models, db, uid, key, 'hr.employee', 'search_read',
                     [[['work_email', '!=', False]]],
                     {'fields': ['name', 'work_email', 'mobile_phone']})

    odoo_map = {}
    for e in odoo_emps:
        for part in e['work_email'].lower().split(';'):
            odoo_map[part.strip()] = e

    enriched = 0
    for emp in employees:
        odoo_e = odoo_map.get(emp['email'])

        # Phone from Odoo mobile_phone
        if not emp['operadora']:
            mobile = odoo_e.get('mobile_phone', '') if odoo_e else ''
            op, num = _parse_odoo_phone(mobile)
            if op and num:
                emp['operadora'] = op
                emp['numero']    = num
                emp['phone_source'] = 'odoo'
                enriched += 1

        # Second names from hard-coded 4-part name map (high confidence only)
        overrides = ODOO_NAME_OVERRIDE.get(emp['email'])
        if overrides:
            seg_nom, seg_ap = overrides
            if not emp['segundo_nombre']:
                emp['segundo_nombre'] = seg_nom
                emp['seg_nom_source'] = 'odoo'
            if not emp['segundo_apellido']:
                emp['segundo_apellido'] = seg_ap
                emp['seg_ap_source']    = 'odoo'

    log.info("Odoo enrichment: %d employees got phone data", enriched)


# ── Email HTML builder ─────────────────────────────────────────────────────────

_ESTADO_CIVIL = {'S': 'Soltero/a', 'C': 'Casado/a', 'D': 'Divorciado/a', 'V': 'Viudo/a'}
_SEXO         = {'M': 'Masculino', 'F': 'Femenino'}

def _field_row(label: str, value, missing: bool, source: str = 'xlsx') -> str:
    """Render one row of the data table."""
    if missing:
        cell = ('<td style="padding:8px 12px;color:#b91c1c;font-weight:600;'
                'background:#fef2f2;letter-spacing:1px;">PENDIENTE</td>')
    else:
        note = (' <span style="font-size:10px;color:#6b7280;">(a confirmar)</span>'
                if source == 'odoo' else '')
        cell = (f'<td style="padding:8px 12px;color:#1e3a5f;">'
                f'{value}{note}</td>')
    return (f'<tr>'
            f'<td style="padding:8px 12px;color:#6b7280;font-size:13px;">{label}</td>'
            f'{cell}'
            f'</tr>')


def _reply_block(emp: dict) -> str:
    """Build the copy-paste reply template for only the missing fields."""
    lines = []
    if not emp['segundo_nombre']:
        lines.append('SEGUNDO NOMBRE        : ________________________________')
    if not emp['segundo_apellido']:
        lines.append('SEGUNDO APELLIDO      : ________________________________')
    if not emp['operadora']:
        lines.append('OPERADORA (3 dígitos) : ____  (ej: 412, 414, 416, 424, 426)')
    if not emp['numero']:
        lines.append('NÚMERO (7 dígitos)    : _______  (ej: 2337463)')
    return '\n'.join(lines)


def _build_html(emp: dict, token: str, dry_run: bool) -> str:
    first    = emp['primer_nombre'].capitalize()
    pronoun  = 'Estimada' if emp['sexo'] == 'F' else 'Estimado'
    form_url = f'{FORM_BASE}/{token}'

    missing_d = not emp['segundo_nombre']
    missing_f = not emp['segundo_apellido']
    missing_o = not emp['operadora']
    missing_p = not emp['numero']
    all_ok    = not any([missing_d, missing_f, missing_o, missing_p])

    phone_display = (f"0{emp['operadora']}-{emp['numero']}"
                     if emp['operadora'] and emp['numero'] else None)

    dry_banner = (
        '<div style="background:#7f1d1d;color:#fef2f2;text-align:center;'
        'padding:10px;font-size:13px;font-weight:bold;letter-spacing:1px;">'
        'MODO PRUEBA — Este correo NO fue enviado al empleado</div>'
    ) if dry_run else ''

    if all_ok:
        intro = ('Por tal razón necesitamos de su amable colaboración de <strong>confirmar</strong> '
                 'que los datos del formulario son correctos haciendo clic en el botón.')
        action_label = 'Confirma que tus datos son correctos usando el enlace a continuación:'
    else:
        missing_labels = []
        if missing_d: missing_labels.append('Segundo Nombre')
        if missing_f: missing_labels.append('Segundo Apellido')
        if missing_o or missing_p: missing_labels.append('Teléfono')
        missing_str = ' · '.join(f'<strong>{m}</strong>' for m in missing_labels)
        intro = (f'Por tal razón necesitamos de su amable colaboración de completar: {missing_str}.<br>'
                 'Haz clic en el botón para ingresar los datos faltantes.')
        action_label = 'Completa los datos faltantes usando el enlace a continuación:'

    # Data table
    table_rows = ''.join([
        _field_row('Número de Cédula', emp['cedula'], not emp['cedula']),
        _field_row('Primer Nombre', emp['primer_nombre'], False),
        _field_row('Segundo Nombre ⚠️' if missing_d else 'Segundo Nombre',
                   emp['segundo_nombre'], missing_d, emp.get('seg_nom_source', 'xlsx')),
        _field_row('Primer Apellido', emp['primer_apellido'], False),
        _field_row('Segundo Apellido ⚠️' if missing_f else 'Segundo Apellido',
                   emp['segundo_apellido'], missing_f, emp.get('seg_ap_source', 'xlsx')),
        _field_row('Fecha de Nacimiento', emp['dob'], not emp['dob']),
        _field_row('Estado Civil', _ESTADO_CIVIL.get(emp['estado_civil'], emp['estado_civil']),
                   not emp['estado_civil']),
        _field_row('Correo Institucional', emp['email'], False),
        _field_row('Teléfono Celular ⚠️' if (missing_o or missing_p) else 'Teléfono Celular',
                   phone_display, missing_o or missing_p,
                   emp.get('phone_source', 'xlsx')),
    ])

    return f'''<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:Arial,sans-serif;">
{dry_banner}
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:30px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0"
       style="background:#ffffff;border-radius:8px;overflow:hidden;
              box-shadow:0 2px 8px rgba(0,0,0,0.1);">

  <tr><td style="background:#1e3a5f;padding:24px 32px;text-align:center;">
    <img src="{LOGO_URL}" height="60" alt="UEIPAB"
         style="border-radius:50%;border:2px solid #ffffff;">
    <p style="color:#ffffff;margin:12px 0 0;font-size:14px;letter-spacing:1px;">
      RECURSOS HUMANOS</p>
  </td></tr>

  <tr><td style="padding:32px;">
    <p style="margin:0 0 16px;font-size:16px;color:#1e3a5f;">
      {pronoun} <strong>{first}</strong>,</p>

    <p style="margin:0 0 16px;color:#374151;line-height:1.6;">
      Estamos en proceso de tener un plan de contingencia y por ende estamos aperturando
      una cuenta nómina adicional en el <strong>Banco Plaza</strong> para todo el personal
      de UEIPAB. {intro}</p>

    <p style="margin:0 0 8px;font-weight:600;color:#1e3a5f;">Tus datos actuales en el sistema:</p>
    <table width="100%" cellpadding="0" cellspacing="0"
           style="border-collapse:collapse;font-size:14px;margin-bottom:0;">
      <thead>
        <tr style="background:#e8f0fe;">
          <th style="padding:8px 12px;text-align:left;color:#1e3a5f;font-size:12px;
                     text-transform:uppercase;">Campo</th>
          <th style="padding:8px 12px;text-align:left;color:#1e3a5f;font-size:12px;
                     text-transform:uppercase;">Valor</th>
        </tr>
      </thead>
      <tbody>{table_rows}</tbody>
    </table>

    <div style="background:#fffbeb;border:2px solid #fbbf24;border-radius:8px;
                padding:20px 24px;margin:20px 0;text-align:center;">
      <p style="margin:0 0 12px;font-size:14px;color:#92400e;font-weight:600;">
        {action_label}
      </p>
      <a href="{form_url}"
         style="display:inline-block;background:#1e3a5f;color:#ffffff;
                text-decoration:none;padding:14px 32px;border-radius:10px;
                font-size:16px;font-weight:bold;letter-spacing:0.3px;">
        Confirmar y/o Actualizar mis Datos &rarr;
      </a>
    </div>

    <hr style="border:none;border-top:1px solid #e5e7eb;margin:20px 0;">
    <p style="margin:0;font-size:13px;color:#6b7280;line-height:1.6;">
      Por favor responde <strong>a la brevedad posible</strong> para no retrasar
      el proceso de apertura de cuentas.<br>
      ¿Dudas? <a href="mailto:recursoshumanos@ueipab.edu.ve"
         style="color:#1e3a5f;">recursoshumanos@ueipab.edu.ve</a>
    </p>
  </td></tr>

  <tr><td style="background:#f9fafb;padding:14px 32px;text-align:center;">
    <p style="margin:0;font-size:11px;color:#9ca3af;">
      Instituto Privado Andrés Bello, CA &middot; El Tigre, Anzoátegui &middot;
      Enlace personal e intransferible para
      {emp['primer_nombre']} {emp['primer_apellido']}.
    </p>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>'''


# ── Send ───────────────────────────────────────────────────────────────────────

def _send_email(models, db, uid, key, emp: dict, token: str, dry_run: bool):
    to_email = CEO_EMAIL if dry_run else emp['email']
    html     = _build_html(emp, token, dry_run)

    missing = sum([
        not emp['segundo_nombre'],
        not emp['segundo_apellido'],
        not emp['operadora'],
        not emp['numero'],
    ])

    mail_id = call(models, db, uid, key, 'mail.mail', 'create', [[{
        'subject':    SUBJECT,
        'body_html':  html,
        'email_to':   to_email,
        'email_from': EMAIL_FROM,
        'reply_to':   REPLY_TO,
        'email_cc':   '' if dry_run else CC_EMAIL,
        'state':      'outgoing',
    }]])

    status = 'OK' if missing == 0 else f'MISSING {missing} field(s)'
    log.info("  mail.mail id=%s → %s  [%s]  %s",
             mail_id, to_email,
             emp['primer_nombre'] + ' ' + emp['primer_apellido'], status)
    return mail_id


def _trigger_queue(models, db, uid, key):
    call(models, db, uid, key, 'ir.cron', 'method_direct_trigger',
         [[MAIL_QUEUE_CRON_ID]])
    log.info("Mail queue cron triggered (id=%s)", MAIL_QUEUE_CRON_ID)


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--live',          action='store_true',
                    help='Send to real employees (all)')
    ap.add_argument('--test-employee', metavar='NAME',
                    help='Send only this employee (partial name match, CEO inbox)')
    ap.add_argument('--send-employee', metavar='NAME',
                    help='Send only this employee to their REAL inbox (partial name match)')
    args = ap.parse_args()

    # --test-employee → dry-run (CEO); --send-employee → live single employee; --live → all live
    dry_run = not (args.live or args.send_employee)

    log.info("Mode: %s%s",
             'DRY-RUN (CEO only)' if dry_run else 'LIVE',
             f' — filter: {args.test_employee or args.send_employee}' if (args.test_employee or args.send_employee) else '')

    employees = _load_xlsx()
    models, db, uid, key = _connect()
    _enrich_from_odoo(employees, models, db, uid, key)

    # Initialize Odoo params + get secret for token generation
    secret = _init_odoo_params(employees, models, db, uid, key)

    # Filter to single employee if requested
    target = employees
    name_filter = args.test_employee or args.send_employee
    if name_filter:
        needle = name_filter.upper()
        target = [e for e in employees
                  if needle in e['primer_nombre'].upper()
                  or needle in e['primer_apellido'].upper()]
        if not target:
            log.error("No employee matching '%s'. Available: %s",
                      name_filter,
                      ', '.join(e['primer_nombre'] for e in employees))
            sys.exit(1)
        log.info("Filtered to %d employee(s): %s",
                 len(target),
                 ', '.join(e['primer_nombre'] + ' ' + e['primer_apellido']
                           for e in target))

    # Summary
    counts = {'ok': 0, 'd': 0, 'f': 0, 'o': 0, 'p': 0}
    for e in employees:
        if not e['segundo_nombre']:   counts['d'] += 1
        if not e['segundo_apellido']: counts['f'] += 1
        if not e['operadora']:        counts['o'] += 1
        if not e['numero']:           counts['p'] += 1
        if all([e['segundo_nombre'], e['segundo_apellido'],
                e['operadora'], e['numero']]):
            counts['ok'] += 1

    log.info("─── Full dataset ───────────────────────────────")
    log.info("  Total: %d | Complete: %d | Miss-D: %d | Miss-F: %d | Miss-phone: %d",
             len(employees), counts['ok'], counts['d'], counts['f'], counts['o'])
    log.info("─── Sending %d email(s) ────────────────────────", len(target))

    sent = 0
    for emp in target:
        token = _compute_token(secret, emp['email'])
        _send_email(models, db, uid, key, emp, token, dry_run)
        sent += 1

    _trigger_queue(models, db, uid, key)
    dest = CEO_EMAIL if dry_run else 'real employee inboxes'
    log.info("Done. %d email(s) → %s", sent, dest)

    if dry_run:
        log.info("Review at %s then run with --live to send to all employees.", CEO_EMAIL)


if __name__ == '__main__':
    main()
