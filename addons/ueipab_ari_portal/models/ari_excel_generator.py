# -*- coding: utf-8 -*-
"""
AR-I Excel Generator Service

Fills the official SENIAT AR-I Excel template with employee data.
"""

import base64
import io
import os
from datetime import date

from odoo import models, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ARIExcelGenerator(models.AbstractModel):
    _name = 'ari.excel.generator'
    _description = 'AR-I Excel Generator Service'

    # Cell mapping for the official SENIAT AR-I template
    CELL_MAP = {
        # Section 1-3: Personal Info
        'name': 'A14',
        'cedula_tipo': 'I14',
        'cedula_numero': 'K14',
        'rif_tipo': 'S14',
        'rif_numero': 'T14',

        # Section 4: Employers
        'empresa_a': 'A18',
        'empresa_b': 'A20',
        'empresa_c': 'G18',
        'empresa_d': 'G20',

        # Section 5: Variation Months
        'var_marzo': 'P19',
        'var_junio': 'S19',
        'var_septiembre': 'U19',
        'var_diciembre': 'W19',

        # Section A: Income
        'income_a': 'C26',
        'income_b': 'C28',
        'income_c': 'L26',
        'income_d': 'L28',
        'total_income': 'T29',
        'total_income_b': 'D32',

        # Section B: UT Value
        'ut_value': 'I32',

        # Section C: Itemized Deductions
        'ded_education': 'T36',
        'ded_insurance': 'T37',
        'ded_medical': 'T38',
        'ded_housing': 'T39',

        # Section E: Unique Deduction
        'ded_unique': 'T45',

        # Section H: Rebajas
        'rebaja_personal': 'K54',
        'rebaja_cargas': 'K55',
        'rebaja_excess': 'C57',

        # Section K: Variation Data
        'ytd_withholding': 'T68',
        'ytd_income': 'T69',

        # Section L: Signature
        'lugar': 'B88',
        'fecha': 'E88',
    }

    @api.model
    def _get_template_path(self):
        """Get the path to the AR-I Excel template."""
        module_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        template_path = os.path.join(
            module_path, 'static', 'templates', 'ARI SENIAT FORMATO.xlsx'
        )
        if not os.path.exists(template_path):
            raise UserError(_(
                'AR-I template not found at: %s\n'
                'Please ensure the SENIAT template is installed.'
            ) % template_path)
        return template_path

    @api.model
    def generate_ari_excel(self, ari_record):
        """
        Generate AR-I Excel file from an hr.employee.ari record.

        Args:
            ari_record: hr.employee.ari record

        Returns:
            tuple: (base64_encoded_data, filename)
        """
        if not openpyxl:
            raise UserError(_(
                'openpyxl library is required to generate AR-I Excel files.\n'
                'Please install it: pip install openpyxl'
            ))

        # Load template
        template_path = self._get_template_path()
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        # Get employee data
        employee = ari_record.employee_id
        contract = ari_record.contract_id
        company = ari_record.company_id or self.env.company

        # --- FILL PERSONAL INFO (Sections 1-3) ---
        ws[self.CELL_MAP['name']] = employee.name or ''

        # Parse cedula
        cedula = employee.identification_id or ''
        cedula_tipo = 'V'
        cedula_numero = ''
        if cedula:
            cedula = cedula.upper().strip()
            if cedula.startswith(('V', 'E', 'P')):
                cedula_tipo = cedula[0]
                cedula_numero = cedula[1:].replace('-', '').replace('.', '')
            else:
                cedula_numero = cedula.replace('-', '').replace('.', '')

        ws[self.CELL_MAP['cedula_tipo']] = cedula_tipo
        try:
            ws[self.CELL_MAP['cedula_numero']] = int(cedula_numero) if cedula_numero else ''
        except ValueError:
            ws[self.CELL_MAP['cedula_numero']] = cedula_numero

        # Parse RIF (if available, otherwise use cedula)
        rif = getattr(employee, 'rif', None) or ''
        if not rif and cedula:
            rif = cedula  # Use cedula as RIF if not specified
        rif_tipo = 'V'
        rif_numero = ''
        if rif:
            rif = rif.upper().strip()
            if rif.startswith(('V', 'E', 'J', 'G', 'P', 'C')):
                rif_tipo = rif[0]
                rif_numero = rif[1:].replace('-', '').replace('.', '')
            else:
                rif_numero = rif.replace('-', '').replace('.', '')

        ws[self.CELL_MAP['rif_tipo']] = rif_tipo
        try:
            ws[self.CELL_MAP['rif_numero']] = int(rif_numero) if rif_numero else ''
        except ValueError:
            ws[self.CELL_MAP['rif_numero']] = rif_numero

        # --- FILL EMPLOYERS (Section 4) ---
        ws[self.CELL_MAP['empresa_a']] = company.name or 'U.E.I.P. AGUSTIN BRICEÑO'
        ws[self.CELL_MAP['empresa_b']] = ari_record.employer_b_name or ''
        ws[self.CELL_MAP['empresa_c']] = ari_record.employer_c_name or ''
        ws[self.CELL_MAP['empresa_d']] = ari_record.employer_d_name or ''

        # --- FILL VARIATION MONTH (Section 5) ---
        ws[self.CELL_MAP['var_marzo']] = ''
        ws[self.CELL_MAP['var_junio']] = ''
        ws[self.CELL_MAP['var_septiembre']] = ''
        ws[self.CELL_MAP['var_diciembre']] = ''

        if ari_record.variation_month == 'march':
            ws[self.CELL_MAP['var_marzo']] = 'X'
        elif ari_record.variation_month == 'june':
            ws[self.CELL_MAP['var_junio']] = 'X'
        elif ari_record.variation_month == 'september':
            ws[self.CELL_MAP['var_septiembre']] = 'X'
        elif ari_record.variation_month == 'december':
            ws[self.CELL_MAP['var_diciembre']] = 'X'

        # --- FILL INCOME (Section A) ---
        ws[self.CELL_MAP['income_a']] = ari_record.income_employer_primary or 0
        ws[self.CELL_MAP['income_b']] = ari_record.income_employer_b or 0
        ws[self.CELL_MAP['income_c']] = ari_record.income_employer_c or 0
        ws[self.CELL_MAP['income_d']] = ari_record.income_employer_d or 0

        # Total income (for cells that aren't formulas in template)
        total_income = ari_record.income_total or 0
        ws[self.CELL_MAP['total_income']] = total_income
        ws[self.CELL_MAP['total_income_b']] = total_income

        # --- FILL UT VALUE (Section B) ---
        ws[self.CELL_MAP['ut_value']] = ari_record.ut_value or 9.00

        # --- FILL DEDUCTIONS (Section C/E) ---
        if ari_record.deduction_type == 'unique':
            # Clear itemized, set unique
            ws[self.CELL_MAP['ded_education']] = ''
            ws[self.CELL_MAP['ded_insurance']] = ''
            ws[self.CELL_MAP['ded_medical']] = ''
            ws[self.CELL_MAP['ded_housing']] = ''
            ws[self.CELL_MAP['ded_unique']] = 774
        else:
            # Itemized deductions
            ws[self.CELL_MAP['ded_education']] = ari_record.deduction_education or ''
            ws[self.CELL_MAP['ded_insurance']] = ari_record.deduction_insurance or ''
            ws[self.CELL_MAP['ded_medical']] = ari_record.deduction_medical or ''
            ws[self.CELL_MAP['ded_housing']] = ari_record.deduction_housing or ''
            ws[self.CELL_MAP['ded_unique']] = ''

        # --- FILL REBAJAS (Section H) ---
        ws[self.CELL_MAP['rebaja_personal']] = 10  # Always 10 UT
        ws[self.CELL_MAP['rebaja_cargas']] = ari_record.cargas_familiares_count * 10
        ws[self.CELL_MAP['rebaja_excess']] = ari_record.rebaja_prior_excess or 0

        # --- FILL VARIATION DATA (Section K) ---
        if ari_record.is_variation and ari_record.ytd_withholding > 0:
            ws[self.CELL_MAP['ytd_withholding']] = ari_record.ytd_withholding
            ws[self.CELL_MAP['ytd_income']] = ari_record.ytd_income

        # --- FILL SIGNATURE (Section L) ---
        ws[self.CELL_MAP['lugar']] = 'CARACAS'  # Default location
        ws[self.CELL_MAP['fecha']] = date.today().strftime('%d/%m/%Y')

        # --- SAVE TO BUFFER ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        cedula_clean = cedula_numero or 'UNKNOWN'
        year = ari_record.fiscal_year or date.today().year
        month = ari_record.variation_month or 'initial'
        filename = f'ARI_{cedula_clean}_{year}_{month}.xlsx'

        # Encode to base64
        excel_data = base64.b64encode(output.read())

        return excel_data, filename

    @api.model
    def preview_calculations(self, ari_record):
        """
        Preview the AR-I calculations without generating Excel.

        Returns a dict with all calculated values for display.
        """
        return {
            'income_total': ari_record.income_total,
            'income_in_ut': ari_record.income_in_ut,
            'deduction_type': ari_record.deduction_type,
            'deductions_in_ut': ari_record.deductions_in_ut if ari_record.deduction_type == 'itemized' else 774,
            'taxable_income_ut': ari_record.taxable_income_ut,
            'estimated_tax_ut': ari_record.estimated_tax_ut,
            'rebajas_total_ut': ari_record.rebajas_total_ut,
            'tax_to_withhold_ut': ari_record.tax_to_withhold_ut,
            'withholding_percentage': ari_record.withholding_percentage,
        }
